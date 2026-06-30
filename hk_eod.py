"""Hong Kong Main Board EOD pipeline. Owns the full HK scan: universe fetch,
Futu data, metrics, strategy filters, dedup, write."""

from __future__ import annotations

import logging
import time
from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import pandas as pd

from futu_sync import get_market_caps_futu
from hk_metrics import build_hk_metrics_cloud

logger = logging.getLogger(__name__)

HKEX_SECURITIES_URL = (
    "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/"
    "ListOfSecurities.xlsx"
)


def fetch_hkex_equities() -> list[str]:
    """Download the HKEX securities xlsx, parse with openpyxl, return Main
    Board equity stock codes as 5-digit strings (e.g., '00700')."""
    from openpyxl import load_workbook

    req = Request(HKEX_SECURITIES_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30) as resp:
        data = resp.read()
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    # Skip 2 leading metadata rows; row 3 has the header.
    next(rows, None)
    next(rows, None)
    header = next(rows, None)
    if header is None:
        return []
    code_idx = next(
        (i for i, h in enumerate(header) if h and "stock code" in str(h).lower()),
        0,
    )
    sub_idx = next(
        (i for i, h in enumerate(header) if h and "sub-category" in str(h).lower()),
        None,
    )

    codes: list[str] = []
    for row in rows:
        if not row or row[code_idx] is None:
            continue
        if sub_idx is not None:
            sub = row[sub_idx]
            if sub != "Equity Securities (Main Board)":
                continue
        raw = str(row[code_idx]).strip()
        if not raw.isdigit():
            continue
        codes.append(raw.zfill(5)[1:])  # 5-digit HKEX code → 4-digit yfinance/TV form ('00700' → '0700')
    return codes


def filter_hk_shorts(
    config: dict, futu_cfg: dict | None = None
) -> tuple[int, list[str]]:
    """Run HK shorts pipeline: fetch HKEX universe, download data via yfinance,
    apply SMA20/volume/cap/dollar-volume/performance/up-days filters.
    Returns (universe_size, filtered_tickers_in_tv_format)."""
    from main import (
        _yf_download_with_retry,
        _get_market_cap,
        _get_closes_volumes,
        _get_ohlc,
        _trim_today,
    )

    logger.info("[HK Shorts] Fetching HKEX equity universe...")
    codes = fetch_hkex_equities()
    logger.info(f"  Found {len(codes)} Main Board equities")

    yf_tickers = [code + ".HK" for code in codes]

    now_hk = datetime.now(ZoneInfo("Asia/Hong_Kong"))
    market_open = 9 <= now_hk.hour < 16 and now_hk.weekday() < 5
    today = now_hk.date()
    if market_open:
        logger.info("  HK market still open, excluding today's incomplete data")

    min_avg_volume = config.get("min_avg_volume", 1_000_000)

    # Phase 1: Download in batches, apply SMA20 +20%, SMA50, and volume filter.
    # Period bumped to 3mo so SMA50 has >= 50 trading days; downstream filters
    # only use up to 22 days so the wider window is harmless.
    # Store per-ticker data to avoid re-downloading later
    logger.info("[HK Shorts] Downloading price data and filtering (this may take several minutes)...")
    batch_size = 500
    phase1 = []
    ticker_closes: dict[str, object] = {}
    ticker_volumes: dict[str, object] = {}
    ticker_highs: dict[str, object] = {}
    ticker_lows: dict[str, object] = {}
    for start in range(0, len(yf_tickers), batch_size):
        batch = yf_tickers[start : start + batch_size]
        logger.info(f"  Batch {start // batch_size + 1}/{(len(yf_tickers) - 1) // batch_size + 1} ({len(batch)} tickers)...")
        batch_data = _yf_download_with_retry(
            batch, period="3mo", progress=False, group_by="ticker", threads=True,
        )
        if batch_data is None or batch_data.empty:
            logger.warning(f"  Batch failed after retries, skipping")
            continue

        single = len(batch) == 1
        for ticker in batch:
            try:
                closes, volumes = _get_closes_volumes(batch_data, ticker, single)
                closes = _trim_today(closes, market_open, today)
                volumes = _trim_today(volumes, market_open, today)

                if len(closes) < 50 or len(volumes) < 20:
                    continue

                last = closes.iloc[-1]
                sma20 = closes.iloc[-20:].mean()
                sma50 = closes.iloc[-50:].mean()
                if (
                    last > sma20 * 1.2
                    and last > sma50
                    and volumes.iloc[-20:].mean() >= min_avg_volume
                ):
                    phase1.append(ticker)
                    ticker_closes[ticker] = closes
                    ticker_volumes[ticker] = volumes
                    highs, lows, _ = _get_ohlc(batch_data, ticker, single)
                    ticker_highs[ticker] = _trim_today(highs, market_open, today)
                    ticker_lows[ticker] = _trim_today(lows, market_open, today)
            except (KeyError, TypeError):
                continue

        if start + batch_size < len(yf_tickers):
            time.sleep(5)

    logger.info(f"  {len(phase1)} after SMA20 +20%, SMA50, and volume filter")
    if not phase1:
        return len(codes), []

    # Phase 2: Market cap. Prefer Futu snapshot (one batch call for all
    # phase1 tickers, real-time `total_market_val` in HKD) over a per-ticker
    # yfinance loop with 0.5s sleep between calls. Fall back to yfinance
    # when OpenD is unreachable or the snapshot call errors.
    min_market_cap = config.get("min_market_cap", 2_000_000_000)
    market_caps: dict[str, float] = {}
    futu_caps = None
    if futu_cfg and futu_cfg.get("enabled"):
        futu_caps = get_market_caps_futu(
            phase1, market="HK",
            host=futu_cfg.get("host", "127.0.0.1"),
            port=futu_cfg.get("port", 11111),
        )
    if futu_caps is not None:
        market_caps = futu_caps
        phase2 = [t for t in phase1 if market_caps.get(t, 0) >= min_market_cap]
        logger.info(
            f"  {len(phase2)} after market cap filter "
            f"(Futu, >= {min_market_cap:,.0f} HKD)"
        )
    else:
        phase2 = []
        for ticker in phase1:
            cap = _get_market_cap(ticker)
            if cap and cap >= min_market_cap:
                phase2.append(ticker)
                market_caps[ticker] = cap
            time.sleep(0.5)
        logger.info(
            f"  {len(phase2)} after market cap filter "
            f"(yfinance, >= {min_market_cap:,.0f} HKD)"
        )
    if not phase2:
        return len(codes), []

    # Phase 3: Dollar volume (price * 20-day avg volume)
    min_dv = config.get("min_dollar_volume", 100_000_000)
    phase3 = []
    for ticker in phase2:
        try:
            closes = ticker_closes[ticker]
            volumes = ticker_volumes[ticker]
            if closes.iloc[-1] * volumes.iloc[-20:].mean() >= min_dv:
                phase3.append(ticker)
        except (KeyError, TypeError):
            continue

    logger.info(f"  {len(phase3)} after dollar volume filter (>= {min_dv:,.0f} HKD)")
    if not phase3:
        return len(codes), []

    # Phase 4: Cap-conditional performance over 2, 3, 4 week windows
    large_cap_thr = config.get("large_cap_threshold", 80_000_000_000)
    mid_cap_thr = config.get("mid_cap_threshold", 16_000_000_000)
    perf_large = config.get("perf_large_cap", 50)
    perf_mid = config.get("perf_mid_cap", 200)
    perf_small = config.get("perf_small_cap", 300)
    perf_weeks = [2, 3, 4]  # trading days: 10, 15, 22

    phase4: set[str] = set()
    for weeks in perf_weeks:
        trading_days = weeks * 5 + (2 if weeks == 4 else 0)  # 10, 15, 22
        week_hits = 0
        for ticker in phase3:
            if ticker in phase4:
                continue
            try:
                closes = ticker_closes[ticker]
                if len(closes) < trading_days + 1:
                    continue
                perf = (closes.iloc[-1] - closes.iloc[-trading_days]) / closes.iloc[-trading_days] * 100
                cap = market_caps[ticker]
                if cap >= large_cap_thr:
                    threshold = perf_large
                elif cap >= mid_cap_thr:
                    threshold = perf_mid
                else:
                    threshold = perf_small
                if perf >= threshold:
                    phase4.add(ticker)
                    week_hits += 1
            except (KeyError, TypeError, ZeroDivisionError):
                continue
        logger.info(f"  {weeks}-week window: {week_hits} new hits")

    logger.info(f"  {len(phase4)} after performance filter (2/3/4 week combined)")
    if not phase4:
        return len(codes), []

    # Phase 4b: ADR% filter
    min_adr = config.get("min_adr_percent", 0)
    adr_days = config.get("adr_days", 20)
    if min_adr > 0:
        adr_passed: set[str] = set()
        for ticker in phase4:
            try:
                highs = ticker_highs[ticker]
                lows = ticker_lows[ticker]
                closes = ticker_closes[ticker]
                n = min(len(highs), len(lows), len(closes), adr_days)
                if n < adr_days:
                    continue
                h = highs.iloc[-adr_days:].values
                l = lows.iloc[-adr_days:].values
                c = closes.iloc[-adr_days:].values
                adr_pct = float(((h - l) / c).mean()) * 100
                if adr_pct >= min_adr:
                    adr_passed.add(ticker)
            except (KeyError, TypeError, ValueError, ZeroDivisionError):
                continue
        phase4 = adr_passed
        logger.info(f"  {len(phase4)} after ADR% filter (>= {min_adr}%, {adr_days}d)")
        if not phase4:
            return len(codes), []

    # Phase 5: Consecutive up days
    min_up_days = config.get("min_consecutive_up_days", 3)
    phase5 = []
    for ticker in phase4:
        try:
            closes = ticker_closes[ticker]
            if len(closes) < 2:
                continue
            consecutive = 0
            for i in range(len(closes) - 1, 0, -1):
                if closes.iloc[i] > closes.iloc[i - 1]:
                    consecutive += 1
                else:
                    break
            if consecutive >= min_up_days:
                phase5.append(ticker)
        except (KeyError, TypeError):
            continue

    logger.info(f"  {len(phase5)} after consecutive up days filter (>= {min_up_days})")

    # Convert to TradingView format: 0700.HK → HKEX:700.
    # TradingView rejects HK tickers with leading zeros (HKEX:0148 won't
    # import — must be HKEX:148); strip leading zeros from the 4-digit
    # yfinance code before prefixing.
    tv_tickers = ["HKEX:" + (t.replace(".HK", "").lstrip("0") or "0") for t in phase5]
    return len(codes), tv_tickers


def fetch_hk_klines_yf(
    codes_4d: list[str],
    period: str = "2y",
    batch_size: int = 500,
) -> dict[str, pd.DataFrame]:
    """Pull daily OHLCV for HK Main Board tickers via yfinance, returning the
    same shape ``build_metrics_frame`` expects: ``{futu_code: DataFrame[
    time_key, open, high, low, close, volume]}``.

    Why yfinance instead of Futu k-line: Futu free/Lv1 tier caps history depth
    for less liquid HK names (~12% of the universe gets ≥ 12 months — see
    PR #7). yfinance reliably gives ≥ 2 years for nearly every Main Board
    listing, so the IBD 12-month RS algorithm actually has a universe to
    rank. Tradeoff: occasional gappy bars on lightly traded names — handled
    via the existing helpers' try/except in main.py.

    Input: 4-digit HK codes from ``fetch_hkex_equities()`` (e.g., ``'0700'``).
    Output keys are Futu format (``'HK.00700'``) so the rest of the pipeline
    (caps re-key, _to_tv, dedup) is unchanged.

    Mirrors the batched download pattern in ``filter_hk_shorts``: 500 tickers
    per batch with ``threads=True``, a 5s pause between batches to be
    polite to yfinance, and ``_trim_today`` of in-progress bars during HK
    market hours. Total runtime ~5–10 min for the full ~2,400 universe.
    """
    if not codes_4d:
        return {}

    # Lazy import to mirror filter_hk_shorts and avoid a circular import
    # (main.py imports from hk_eod, hk_eod imports from main only at call time).
    from main import (
        _yf_download_with_retry,
        _retry_sparse_in_batch,
        _trim_today,
    )

    yf_tickers = [f"{code}.HK" for code in codes_4d]
    now_hk = datetime.now(ZoneInfo("Asia/Hong_Kong"))
    market_open = 9 <= now_hk.hour < 16 and now_hk.weekday() < 5
    today = now_hk.date()
    if market_open:
        logger.info("[HK Longs] HK market in session — _trim_today will drop today's incomplete bar.")

    result: dict[str, pd.DataFrame] = {}
    drop_buckets = {"missing": 0, "sparse": 0, "shape_err": 0, "ok": 0}
    n_batches = (len(yf_tickers) - 1) // batch_size + 1
    for bidx, start in enumerate(range(0, len(yf_tickers), batch_size), start=1):
        batch = yf_tickers[start:start + batch_size]
        logger.info(f"[HK Longs] yfinance batch {bidx}/{n_batches} ({len(batch)} tickers)...")
        batch_data = _yf_download_with_retry(
            batch, period=period, progress=False, group_by="ticker", threads=True,
        )
        if batch_data is None or batch_data.empty:
            logger.warning(f"[HK Longs]   batch failed after retries, skipping {len(batch)} tickers")
            drop_buckets["missing"] += len(batch)
            continue

        # yfinance batch responses occasionally drop legitimate tickers (all-
        # NaN columns) due to transient Yahoo errors; the single-ticker retry
        # usually recovers them. Without this we lose ~30-40% of the universe
        # on a flaky-yfinance day (e.g. 1384/2420 = 57% on 2026-05-06 13:57).
        _retry_sparse_in_batch(batch_data, batch, period=period, min_rows=20)

        # Build OHLCV per ticker by reading directly from the batch_data
        # MultiIndex (rather than calling _get_closes_volumes / _get_ohlc
        # which each do their own .dropna() and risk shape mismatch).
        for ticker in batch:
            try:
                if ticker not in batch_data.columns.get_level_values(0).unique():
                    drop_buckets["missing"] += 1
                    continue
                sub = batch_data[ticker]
                df = pd.DataFrame({
                    "time_key": pd.to_datetime(sub.index),
                    "open": sub["Open"].values,
                    "high": sub["High"].values,
                    "low": sub["Low"].values,
                    "close": sub["Close"].values,
                    "volume": sub["Volume"].values,
                })
                df = df.dropna(subset=["close"]).reset_index(drop=True)
                # Trim today's incomplete bar during HK market hours
                if market_open and len(df) and df["time_key"].iloc[-1].date() == today:
                    df = df.iloc[:-1].reset_index(drop=True)
                if len(df) < 2:
                    drop_buckets["sparse"] += 1
                    continue
                # 4-digit yfinance ticker '0700.HK' → 5-digit Futu code 'HK.00700'
                code_4d_only = ticker.replace(".HK", "")
                futu_code = f"HK.0{code_4d_only}"
                result[futu_code] = df
                drop_buckets["ok"] += 1
            except (KeyError, TypeError, ValueError) as e:
                drop_buckets["shape_err"] += 1
                if drop_buckets["shape_err"] <= 5:
                    logger.warning(f"[HK Longs]   {ticker}: {type(e).__name__}: {e}")
                continue

        if bidx < n_batches:
            time.sleep(5)

    logger.info(
        f"[HK Longs] yfinance fetch done: {len(result)}/{len(yf_tickers)} tickers "
        f"(ok={drop_buckets['ok']}, missing={drop_buckets['missing']}, "
        f"sparse={drop_buckets['sparse']}, shape_err={drop_buckets['shape_err']})"
    )
    return result


def fetch_hsi_kline_yf(period: str = "2y") -> pd.DataFrame | None:
    """Fetch HSI daily k-line via yfinance (`^HSI`). Returns DataFrame with
    the same columns as ``fetch_hk_klines_yf`` rows, or ``None`` on failure.

    Note: yfinance with ``group_by="ticker"`` returns a MultiIndex DataFrame
    even when the input is a single-element list, so we index via
    ``data["^HSI"]`` to get the OHLCV grid directly.
    """
    from main import _yf_download_with_retry

    data = _yf_download_with_retry(
        ["^HSI"], period=period, progress=False, group_by="ticker", threads=False,
    )
    if data is None or data.empty:
        logger.warning("[HK Longs] HSI yfinance fetch returned empty")
        return None
    try:
        sub = data["^HSI"]
        df = pd.DataFrame({
            "time_key": pd.to_datetime(sub.index),
            "open": sub["Open"].values,
            "high": sub["High"].values,
            "low": sub["Low"].values,
            "close": sub["Close"].values,
            "volume": sub["Volume"].values,
        })
        return df.dropna(subset=["close"]).reset_index(drop=True)
    except (KeyError, TypeError, ValueError) as e:
        logger.warning(f"[HK Longs] HSI parse failed: {e}")
        return None


def build_metrics_frame(
    klines: dict[str, pd.DataFrame],
    market_caps: dict[str, float],
) -> pd.DataFrame:
    """Reduce a {code: kline_df} dict + caps to a metrics DataFrame indexed by
    code. Tickers without enough history for a given metric get NaN/False.

    Columns:
      n_rows, market_cap, last_price, prev_close, gap_pct, rvol, avg_vol_20d,
      avg_dollar_vol_20d, adr_pct, sma50, sma200, above_sma50, above_sma200,
      perf_4w, perf_13w, perf_26w, perf_ytd, perf_52w, consecutive_up_days

    ``n_rows`` is the per-ticker k-line row count — it is published on the
    cloud metrics CSV (not in the drop list) so the HK IPO sidecar can bucket
    by history depth on both the cloud and local-fallback paths.
    """
    rows: list[dict] = []
    today_year = pd.Timestamp.today().year
    for code, df in klines.items():
        if df is None or df.empty or len(df) < 2:
            continue
        closes = df["close"].astype(float).values
        volumes = df["volume"].astype(float).values
        highs = df["high"].astype(float).values
        lows = df["low"].astype(float).values
        times = df["time_key"]
        n = len(closes)

        last = float(closes[-1])
        prev = float(closes[-2])
        gap = (last - prev) / prev * 100.0 if prev > 0 else float("nan")
        avg_vol_20 = float(volumes[-20:].mean()) if n >= 20 else float("nan")
        rvol = (
            float(volumes[-1] / volumes[-21:-1].mean())
            if n >= 21 and volumes[-21:-1].mean() > 0
            else float("nan")
        )
        avg_dv_20 = last * avg_vol_20 if n >= 20 else float("nan")

        if n >= 20:
            adr = float(((highs[-20:] - lows[-20:]) / closes[-20:]).mean()) * 100
        else:
            adr = float("nan")

        sma50 = float(closes[-50:].mean()) if n >= 50 else float("nan")
        sma200 = float(closes[-200:].mean()) if n >= 200 else float("nan")
        above_sma50 = bool(n >= 50 and last > sma50)
        above_sma200 = bool(n >= 200 and last > sma200)

        def _perf(days: int) -> float:
            if n <= days:
                return float("nan")
            past = closes[-days - 1]
            return (last - past) / past * 100.0 if past > 0 else float("nan")

        perf_4w = _perf(20)
        perf_13w = _perf(65)
        perf_26w = _perf(130)
        perf_52w = _perf(252) if n > 252 else float("nan")

        # YTD: find earliest close in the current year
        ytd_mask = times.dt.year == today_year
        if ytd_mask.any():
            first_ytd = float(closes[ytd_mask.values][0])
            perf_ytd = (last - first_ytd) / first_ytd * 100.0 if first_ytd > 0 else float("nan")
        else:
            perf_ytd = float("nan")

        # Consecutive up days from the tail
        cu = 0
        for i in range(n - 1, 0, -1):
            if closes[i] > closes[i - 1]:
                cu += 1
            else:
                break

        rows.append({
            "code": code,
            "n_rows": n,
            "market_cap": market_caps.get(code, float("nan")),
            "last_price": last,
            "prev_close": prev,
            "gap_pct": gap,
            "rvol": rvol,
            "avg_vol_20d": avg_vol_20,
            "avg_dollar_vol_20d": avg_dv_20,
            "adr_pct": adr,
            "sma50": sma50,
            "sma200": sma200,
            "above_sma50": above_sma50,
            "above_sma200": above_sma200,
            "perf_4w": perf_4w,
            "perf_13w": perf_13w,
            "perf_26w": perf_26w,
            "perf_ytd": perf_ytd,
            "perf_52w": perf_52w,
            "consecutive_up_days": cu,
        })
    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows).set_index("code")
    # Keep above_* as Python bool singletons (not numpy.bool_) so that
    # ``row["above_sma50"] is False`` works correctly in callers/tests.
    for col in ("above_sma50", "above_sma200"):
        if col in result.columns:
            result[col] = result[col].astype(object)
    return result


HK_STRATEGY_PRIORITY = ["EarningsGap", "HighVolume", "GapUp", "Leaders", "RS"]


def apply_strategy_filters(
    metrics: pd.DataFrame,
    settings: dict,
    longs_cfg: list[dict],
    leaders_cfg: list[dict],
    rs_enabled: bool,
) -> dict[str, list[str]]:
    """Apply every strategy gate against the metrics frame and return a dict
    of {strategy_name: [code, ...]}. Codes are still in Futu format
    (``HK.00700``); caller converts to TradingView ``HKEX:00700`` later.

    Strategies returned: EarningsGap, HighVolume, GapUp, Leaders, RS
    (the last is always returned but empty when rs_enabled is False).
    """
    if metrics.empty:
        return {s: [] for s in HK_STRATEGY_PRIORITY}

    cap = settings.get("min_market_cap", 300_000_000)
    dvol = settings.get("min_dollar_volume", 100_000_000)
    avg_vol = settings.get("min_avg_volume", 500_000)
    adr = settings.get("min_adr_percent", 4.0)
    price = settings.get("min_price", 20.0)

    # --- Funnel diagnostics: log how many tickers survive each baseline gate
    # individually. Without this, a "0 candidates" outcome could be caused by
    # any of cap/avg_vol/$vol/adr/price/SMA/RS — operator can't tell where
    # the filter is too tight without per-stage counts.
    #
    # SMA50 & SMA200 are part of the universal baseline (mirrors US Longs
    # configs, which all carry ta_sma50_pa + ta_sma200_pa). Every HK long-
    # side strategy gates on the same trend filter.
    n = len(metrics)
    cap_ok = (metrics["market_cap"] >= cap)
    vol_ok = (metrics["avg_vol_20d"] >= avg_vol)
    dvol_ok = (metrics["avg_dollar_vol_20d"] >= dvol)
    adr_ok = (metrics["adr_pct"] >= adr)
    price_ok = (metrics["last_price"] >= price)
    sma_ok = metrics["above_sma50"].astype(bool) & metrics["above_sma200"].astype(bool)
    base = cap_ok & vol_ok & dvol_ok & adr_ok & price_ok & sma_ok
    logger.info(
        f"[HK Longs] baseline funnel (n={n}): "
        f"cap>={cap:,.0f} {int(cap_ok.sum())}; "
        f"avg_vol>={avg_vol:,.0f} {int(vol_ok.sum())}; "
        f"$vol>={dvol:,.0f} {int(dvol_ok.sum())}; "
        f"ADR>={adr}% {int(adr_ok.sum())}; "
        f"price>={price} {int(price_ok.sum())}; "
        f"above SMA50 & SMA200 {int(sma_ok.sum())}; "
        f"baseline-AND {int(base.sum())}"
    )

    # Per-strategy parameter lookup
    by_key = {item.get("key"): item for item in longs_cfg}
    eg = by_key.get("earnings_gap", {})
    hv = by_key.get("high_volume", {})
    gu = by_key.get("gap_up", {})

    eg_min_rvol = float(eg.get("min_relative_volume", 3))
    eg_min_gap = float(eg.get("min_gap_percent", 3.0))
    hv_min_rvol = float(hv.get("min_relative_volume", 3))
    gu_min_gap = float(gu.get("min_gap_percent", 5.0))

    earnings_gap_mask = base & (metrics["gap_pct"] >= eg_min_gap) & (metrics["rvol"] >= eg_min_rvol)
    high_volume_mask = base & (metrics["rvol"] >= hv_min_rvol)
    gap_up_mask = base & (metrics["gap_pct"] >= gu_min_gap)

    # Leaders: baseline (incl. SMA50 & SMA200) + any of the perf windows
    perf_any = (
        (metrics["perf_4w"] >= _leader_threshold(leaders_cfg, "min_perf_4w"))
        | (metrics["perf_13w"] >= _leader_threshold(leaders_cfg, "min_perf_13w"))
        | (metrics["perf_26w"] >= _leader_threshold(leaders_cfg, "min_perf_26w"))
        | (metrics["perf_ytd"] >= _leader_threshold(leaders_cfg, "min_perf_ytd"))
        | (metrics["perf_52w"] >= _leader_threshold(leaders_cfg, "min_perf_52w"))
    ).fillna(False)

    leaders_mask = base & perf_any
    # RS group: baseline (incl. SMA50/200), no perf window. Always computed,
    # caller decides whether to actually emit it based on HSI trigger.
    rs_mask = base

    # Per-strategy diagnostic — show pre-RS-gate counts so operator can see
    # which strategy had no setups today vs which got cut by a downstream gate.
    logger.info(
        f"[HK Longs] strategy funnel (pre-RS-gate): "
        f"EarningsGap (gap>={eg_min_gap}% AND rvol>={eg_min_rvol}) {int(earnings_gap_mask.sum())}; "
        f"HighVolume (rvol>={hv_min_rvol}) {int(high_volume_mask.sum())}; "
        f"GapUp (gap>={gu_min_gap}%) {int(gap_up_mask.sum())}; "
        f"Leaders (+SMA & perf) {int(leaders_mask.sum())}; "
        f"RS (+SMA) {int(rs_mask.sum())}"
    )

    return {
        "EarningsGap": metrics.index[earnings_gap_mask].tolist(),
        "HighVolume": metrics.index[high_volume_mask].tolist(),
        "GapUp": metrics.index[gap_up_mask].tolist(),
        "Leaders": metrics.index[leaders_mask].tolist(),
        "RS": metrics.index[rs_mask].tolist() if rs_enabled else [],
    }


def _leader_threshold(leaders_cfg: list[dict], key: str) -> float:
    """Find the threshold for a given perf window across the [[hk_leaders]]
    list. Returns +inf if no matching entry (so the OR short-circuits)."""
    for item in leaders_cfg:
        if key in item:
            return float(item[key])
    return float("inf")


def _rs_line_group_note(codes, rs_line_tbl) -> str:
    """' | RS↓N' where N = how many of `codes` have their RS line below its MA.
    Empty string when the table is absent or no code is below — keeps the
    summary line clean. `codes` are Futu codes (rs_line_tbl index format)."""
    if rs_line_tbl is None or getattr(rs_line_tbl, "empty", True):
        return ""
    if "rs_below_ma" not in rs_line_tbl.columns:
        return ""
    n = 0
    for c in codes:
        if c in rs_line_tbl.index:
            v = rs_line_tbl.loc[c, "rs_below_ma"]
            if not pd.isna(v) and int(v) == 1:
                n += 1
    return f" | RS↓{n}" if n else ""


def dedup_by_priority(
    raw: dict[str, list[str]],
    priority: list[str] | None = None,
) -> dict[str, list[str]]:
    """Given {strategy: [code,...]}, walk in priority order and drop codes
    from later strategies that already appeared in an earlier one. Default
    priority: EarningsGap > HighVolume > GapUp > Leaders > RS."""
    order = priority or HK_STRATEGY_PRIORITY
    seen: set[str] = set()
    out: dict[str, list[str]] = {}
    for name in order:
        codes = [c for c in raw.get(name, []) if c not in seen]
        out[name] = codes
        seen.update(codes)
    return out


def filter_hk_ipo_candidates(
    metrics: pd.DataFrame,
    rs_table_3m: pd.DataFrame | None,
    hk_settings: dict,
) -> tuple[list[str], dict[str, int]]:
    """筛选 HK IPO 候选 (< 253 行历史的 HKEX 主板 ticker)。

    历史深度现在取自 ``metrics['n_rows']``，不再接收 klines —— 直接遍历
    metrics 帧。这样云端 (data/hk_metrics/<date>.csv) 与本地回退两条路径
    都能跑 IPO（n_rows 随 metrics 一起发布/回填）。`'no_metrics'` 桶已移除：
    既然直接遍历 metrics，不可能出现 code 不在 metrics 里的情况。

    过滤阶梯（同一 ticker 按顺序走，命中任一即 drop 并计数）：
      - n_rows < 20           → drops['min_history']     (day-20 floor)
      - cap < min_market_cap  → drops['cap']
      - price < min_price     → drops['price']
      - if has 20-day metrics:
          avg_vol < min_avg_volume       → drops['avg_vol']
          $vol   < min_dollar_volume     → drops['dvol']
          ADR%   < min_adr_percent       → drops['adr']
      - if has SMA50: not above SMA50    → drops['sma50']
      - if has SMA200: not above SMA200  → drops['sma200']
      - if n_rows >= 64 and rs_table_3m is not None:        (3M RS 闸门)
          ticker not in rs_table_3m      → drops['rs_3m_missing']
          rs_percentile < threshold      → drops['rs_3m']

    253 行及以上的 ticker 不在此函数范围（走长线流水线），直接跳过。

    Returns:
        (kept_codes, drop_counts) — kept_codes 是 Futu 5-digit 格式
        (例如 "HK.00700")；drop_counts 包含上述所有 reason keys，
        每个 key 至少为 0。
    """
    ipo_cap = hk_settings.get("min_market_cap", 300_000_000)
    ipo_min_price = hk_settings.get("min_price", 20.0)
    ipo_min_avg_vol = hk_settings.get("min_avg_volume", 500_000)
    ipo_min_dvol = hk_settings.get("min_dollar_volume", 100_000_000)
    ipo_min_adr = hk_settings.get("min_adr_percent", 3.5)
    rs_3m_threshold = int(hk_settings.get("min_rs_percentile_longs_3m", 90))

    kept: list[str] = []
    drops: dict[str, int] = {
        "min_history": 0,
        "cap": 0, "price": 0,
        "avg_vol": 0, "dvol": 0, "adr": 0,
        "sma50": 0, "sma200": 0,
        "rs_3m": 0, "rs_3m_missing": 0,
    }

    for code in metrics.index:
        row = metrics.loc[code]
        n = int(row["n_rows"])
        if n >= 253:
            continue  # 完整历史 — 由长线流水线处理
        if n < 20:
            drops["min_history"] += 1
            continue
        if not (pd.notna(row["market_cap"]) and row["market_cap"] >= ipo_cap):
            drops["cap"] += 1
            continue
        if not (pd.notna(row["last_price"]) and row["last_price"] >= ipo_min_price):
            drops["price"] += 1
            continue
        # 20-day metrics: guaranteed non-NaN since n_rows >= 20 (checked above).
        if pd.notna(row["avg_vol_20d"]) and row["avg_vol_20d"] < ipo_min_avg_vol:
            drops["avg_vol"] += 1
            continue
        if pd.notna(row["avg_dollar_vol_20d"]) and row["avg_dollar_vol_20d"] < ipo_min_dvol:
            drops["dvol"] += 1
            continue
        if pd.notna(row["adr_pct"]) and row["adr_pct"] < ipo_min_adr:
            drops["adr"] += 1
            continue
        # 50/200-day conditional — sma50/sma200 为 NaN 时不查 above_sma
        if pd.notna(row["sma50"]) and not bool(row["above_sma50"]):
            drops["sma50"] += 1
            continue
        if pd.notna(row["sma200"]) and not bool(row["above_sma200"]):
            drops["sma200"] += 1
            continue
        # 3M RS gate — 仅当 n_rows >= 64 (即 3M RS 算法可计算)、表存在、且
        # 阈值 > 0 时触发。threshold=0 关闭整个闸门 (与 filter_by_rs 行为一致)。
        if (
            n >= 64
            and rs_3m_threshold > 0
            and rs_table_3m is not None
            and not rs_table_3m.empty
        ):
            if code not in rs_table_3m.index:
                drops["rs_3m_missing"] += 1
                continue
            if int(rs_table_3m.loc[code, "rs_percentile"]) < rs_3m_threshold:
                drops["rs_3m"] += 1
                continue
        kept.append(code)

    return kept, drops


def hsi_day_change_pct(
    host: str = "127.0.0.1", port: int = 11111
) -> float | None:
    """Return today's HSI day change in percent, derived from
    ``(last_price - prev_close_price) / prev_close_price * 100``. Uses
    Futu code ``HK.800000`` for HSI. Returns ``None`` on any failure
    (caller should treat None as 'trigger condition not met')."""
    try:
        from futu import OpenQuoteContext, RET_OK
    except ImportError:
        return None

    from futu_sync import _opend_reachable
    if not _opend_reachable(host, port):
        return None

    ctx = None
    try:
        ctx = OpenQuoteContext(host=host, port=port)
        ret, data = ctx.get_market_snapshot(["HK.800000"])
        if ret != RET_OK or data is None or data.empty:
            return None
        row = data.iloc[0]
        last = float(row.get("last_price") or 0)
        prev = float(row.get("prev_close_price") or 0)
        if prev <= 0:
            return None
        return (last - prev) / prev * 100.0
    except Exception:
        return None
    finally:
        if ctx is not None:
            try:
                ctx.close()
            except Exception:
                pass


from pathlib import Path


def _to_tv(code: str) -> str:
    """``HK.00700`` → ``HKEX:700``. TradingView rejects HK tickers with
    leading zeros (e.g. ``HKEX:0148`` won't import — must be ``HKEX:148``).
    Strip the ``HK.`` prefix and any leading zeros; fall back to ``"0"``
    in the degenerate all-zero case so the output is never empty."""
    stripped = code.replace("HK.", "", 1).lstrip("0")
    return "HKEX:" + (stripped or "0")


def _tv_to_code(tv: str) -> str:
    """``HKEX:700`` → ``HK.00700`` — inverse of ``_to_tv``, re-padding to the
    5-digit Futu code used as the RS-table index. ``HKEX:1`` → ``HK.00001``."""
    num = tv.replace("HKEX:", "", 1)
    return "HK." + num.zfill(5)


def run_hk_eod(
    config: dict,
    output_dir: Path,
    today_iso: str,
    write_watchlist,         # callable from main.py
    write_webull,            # callable from main.py (_write_webull)
    futu_sync,               # callable from main.py (_futu_sync)
    tv_sync,                 # callable from main.py (_tv_sync)
    load_seen,               # callable from main.py (_load_seen)
    persist_seen,            # callable from main.py (_persist_seen)
    eod_seen_path,           # callable from main.py (_eod_seen_path)
    dedup_seen,              # callable from main.py (_dedup_seen)
) -> None:
    """Top-level HK EOD pipeline. Runs the existing HK Shorts (unchanged)
    plus the new long-side strategies (EarningsGap, HighVolume, GapUp,
    Leaders, conditional RS). Writes 6 dated .txt files, mirrors to Webull,
    and Futu-syncs each one."""
    hk_settings = config.get("hk_settings") or {}
    hk_longs = config.get("hk_longs") or []
    hk_leaders = config.get("hk_leaders") or []
    futu_cfg = config.get("futu") or {}
    hk_output_dir = output_dir / "TV" / "HK"
    hk_output_dir.mkdir(parents=True, exist_ok=True)

    fmt = config.get("settings", {}).get("output_format", "comma")
    host = futu_cfg.get("host", "127.0.0.1")
    port = int(futu_cfg.get("port", 11111))

    # --- HK Shorts (existing pipeline, unchanged) ---
    hk_shorts_cfg = config.get("hk_shorts")
    if hk_shorts_cfg:
        hk_shorts_cfg.setdefault(
            "min_adr_percent",
            config.get("settings", {}).get("min_adr_percent", 4.0),
        )
        hk_shorts_cfg.setdefault(
            "adr_days", config.get("settings", {}).get("adr_days", 20)
        )
        try:
            total, hk_shorts_tv = filter_hk_shorts(hk_shorts_cfg, futu_cfg=futu_cfg)
            sorted_hk = sorted(hk_shorts_tv)
            dated = hk_output_dir / f"{today_iso}_Shorts.txt"
            write_watchlist(sorted_hk, dated, fmt)
            logger.info(f"[HK Shorts] {len(sorted_hk)} -> {dated}")
            write_webull(sorted_hk, dated, output_dir)
            futu_sync(config, "hk_shorts", sorted_hk, "HK")
            tv_sync(config, "hk_shorts", sorted_hk, "HK")
        except Exception as e:
            logger.warning(f"[HK Shorts] Failed: {e}")

    # --- HK Long-side ---
    # Data source for the fallback path: yfinance for k-line + HSI (deeper
    # history than Futu Lv1 gives for less liquid HK names — see PR #7
    # diagnostic that found only 282/2400 made it past the 12-month threshold
    # via Futu). Futu is still used for market caps + the live HSI day-change
    # snapshot. The cloud-first path skips the local yfinance download entirely.

    # Cloud-first: the RS workflow already fetched every HK k-line; it also
    # publishes a metrics frame (data/hk_metrics/<date>.csv). Pull that to
    # skip the local ~2,400-ticker yfinance download (which throttles to
    # ~66% coverage). On any miss, fall back to the local live fetch — a
    # stale frame would carry wrong gap_pct/rvol, so there is no walk-back.
    today_d = date.today()
    metrics = build_hk_metrics_cloud(output_dir, today_d)

    if metrics is not None:
        logger.info(
            f"[HK Longs] Using cloud metrics: {len(metrics)} tickers "
            "(local k-line fetch skipped)"
        )
        # Cloud metrics are always settled-close (published at 19:00 HKT);
        # the pre-20:00 bar-trim is not needed and use_yesterday stays False.
        use_yesterday = False
        logger.info("[HK Longs] Fetching market caps via Futu snapshot...")
        tv_codes = [_to_tv(c) for c in metrics.index]
        futu_caps_by_tv = (
            get_market_caps_futu(tv_codes, market="HK", host=host, port=port) or {}
        )
        caps = {
            f"HK.{tv.replace('HKEX:', '').zfill(5)}": v
            for tv, v in futu_caps_by_tv.items()
        }
        metrics["market_cap"] = metrics.index.map(lambda c: caps.get(c, float("nan")))
    else:
        logger.warning(
            "[HK Longs] Cloud metrics unavailable; falling back to local "
            "yfinance fetch (throttle-prone, partial coverage)"
        )
        logger.info("[HK Longs] Fetching universe...")
        codes_4d = fetch_hkex_equities()
        logger.info(f"  Universe: {len(codes_4d)} codes")

        logger.info("[HK Longs] Fetching daily OHLCV via yfinance (~5-10 min)...")
        klines = fetch_hk_klines_yf(codes_4d, period="2y")

        if klines:
            lens = sorted(len(df) for df in klines.values())
            n = len(lens)
            ge253 = sum(1 for l in lens if l >= 253)
            ge200 = sum(1 for l in lens if l >= 200)
            ge100 = sum(1 for l in lens if l >= 100)
            ge20 = sum(1 for l in lens if l >= 20)
            median = lens[n // 2] if n else 0
            logger.info(
                f"[HK Longs] k-line depth: n={n}, median={median} rows; "
                f">=253: {ge253} ({100*ge253//max(n,1)}%), >=200: {ge200}, "
                f">=100: {ge100}, >=20: {ge20}"
            )

        # Trim incomplete today's bar if running before 20:00 HKT (only the
        # local-fetch path needs this — cloud metrics are always settled-close).
        hkt_now = datetime.now(ZoneInfo("Asia/Hong_Kong"))
        use_yesterday = hkt_now.hour < 20
        if use_yesterday and klines:
            today_d_local = hkt_now.date()
            trimmed = 0
            for code, df in list(klines.items()):
                mask = df["time_key"].dt.date < today_d_local
                if (~mask).any():
                    trimmed += 1
                klines[code] = df[mask].reset_index(drop=True)
            logger.info(
                f"[HK Longs] Pre-20:00 HKT run (now {hkt_now.strftime('%H:%M')}); "
                f"trimmed today's incomplete bar from {trimmed} tickers — "
                f"using previous-day close as 'latest'."
            )

        logger.info("[HK Longs] Fetching market caps via Futu snapshot...")
        tv_codes = [_to_tv(c) for c in klines.keys()]
        futu_caps_by_tv = (
            get_market_caps_futu(tv_codes, market="HK", host=host, port=port) or {}
        )
        caps = {
            f"HK.{tv.replace('HKEX:', '').zfill(5)}": v
            for tv, v in futu_caps_by_tv.items()
        }

        logger.info("[HK Longs] Building metrics frame...")
        metrics = build_metrics_frame(klines, caps)

    logger.info(f"  Metrics: {len(metrics)} tickers with usable history")

    # --- RS tables (12M + 3M) ---
    # 双闸门: 先过 12M RS >= threshold, 再过 3M RS >= threshold_3m.
    # 两张表由 GitHub Actions 在新 IP 上跑全宇宙 (.github/workflows/update_hk_rs.yml)
    # 并发布到 data/hk_rs/<date>.csv; 本地只做 HTTP 拉取 + 3 天陈旧回退, 命中后
    # 镜像到 hk_rs_rating_<date>.csv / hk_rs_rating_3m_<date>.csv 供当天重跑短路。
    # 计算搬到云端是因为家用 IP 抓 ~2400 港股会被 yfinance 限流 (2026-05-25 仅
    # ~50% 覆盖), 让百分位分布只建立在半个宇宙上。"missing -> passthrough" 策略
    # 两层都遵守; 任一阈值设为 0 即关闭该层 (filter_by_rs 内部短路)。
    from hk_rs import filter_by_rs, build_hk_rs_tables
    rs_table_12m, rs_table_3m, rs_line_tbl = build_hk_rs_tables(output_dir, today_d)

    # --- Apply per-strategy filters ---
    # The conditional RS group keys off HSI's "today" day-change. When
    # use_yesterday is True the live HSI snapshot reflects a state that
    # doesn't match the trimmed k-line data, so the trigger is meaningless
    # — skip the RS group in that case.
    rs_trigger = hk_settings.get("hsi_rs_trigger", -1.2)
    if use_yesterday:
        hsi_change = None
        rs_enabled = False
        logger.info(
            "[HK Longs] Pre-20:00 run uses yesterday's close — HSI conditional "
            "RS group skipped (live HSI day-change does not match trimmed bars)."
        )
    else:
        hsi_change = hsi_day_change_pct(host=host, port=port)
        rs_enabled = hsi_change is not None and hsi_change <= rs_trigger
        logger.info(
            f"[HK Longs] HSI day-change={hsi_change} (trigger {rs_trigger}); "
            f"RS group {'ENABLED' if rs_enabled else 'skipped'}"
        )

    raw = apply_strategy_filters(metrics, hk_settings, hk_longs, hk_leaders, rs_enabled)

    # --- RS double gate (12M ∩ 3M) ---
    threshold_12m = int(hk_settings.get("min_rs_percentile_longs", 90))
    threshold_3m  = int(hk_settings.get("min_rs_percentile_longs_3m", 90))
    pre_counts   = {n: len(c) for n, c in raw.items()}
    after_12m    = {n: filter_by_rs(c, rs_table_12m, threshold_12m) for n, c in raw.items()}
    after_3m     = {n: filter_by_rs(c, rs_table_3m,  threshold_3m)  for n, c in after_12m.items()}
    raw = after_3m
    logger.info(
        f"[HK Longs] RS 12m>={threshold_12m} ∩ 3m>={threshold_3m}: "
        + ", ".join(
            f"{n} {pre_counts[n]}→{len(after_12m[n])}→{len(after_3m[n])}"
            for n in HK_STRATEGY_PRIORITY
        )
    )
    post_rs_counts = {n: len(c) for n, c in raw.items()}

    # --- Within-day cross-strategy priority dedup ---
    # RS is the conditional weak-market scan; the entire point is to
    # re-surface names already collected in other groups when HSI dumps.
    # So priority dedup is restricted to the 4 first-sighting strategies,
    # and RS is spliced back in untouched.
    dedup = dedup_by_priority(
        raw,
        priority=["EarningsGap", "HighVolume", "GapUp", "Leaders"],
    )
    dedup["RS"] = list(raw.get("RS", []))
    logger.info(
        "[HK Longs] within-day priority dedup: "
        + ", ".join(f"{n} {post_rs_counts[n]}→{len(dedup[n])}" for n in HK_STRATEGY_PRIORITY)
    )

    # --- Cross-day master dedup ---
    # RS is excluded — see comment on within-day priority dedup above.
    # An RS hit must NOT consult or mutate the long-side master, because
    # the whole point of the conditional scan is to re-detect strong
    # names on weak-market days regardless of prior sightings.
    seen_path = eod_seen_path(output_dir, "HK")
    seen = load_seen(seen_path)
    final: dict[str, list[str]] = {}
    for name, codes_list in dedup.items():
        tv = sorted(_to_tv(c) for c in codes_list)
        if name == "RS":
            final[name] = tv
            continue
        tag = f"[HK {name}]"
        tv = dedup_seen(tag, tv, seen, seen_path)
        final[name] = tv

    # --- Write outputs + Futu sync ---
    futu_key = {
        "EarningsGap": "hk_longs_earnings_gap",
        "HighVolume":  "hk_longs_high_volume",
        "GapUp":       "hk_longs_gap_up",
        "Leaders":     "hk_leaders",
        "RS":          "hk_rs",
    }
    for name, tv in final.items():
        dated = hk_output_dir / f"{today_iso}_{name}.txt"
        write_watchlist(tv, dated, fmt)
        logger.info(f"[HK {name}] {len(tv)} -> {dated}")
        write_webull(tv, dated, output_dir)
        futu_sync(config, futu_key[name], tv, "HK")
        tv_sync(config, futu_key[name], tv, "HK")

    # --- RS New High strong sub-list (HK) ---
    # Mirror of the US RS-NH block: of today's long-side survivors (the 4
    # first-sighting groups incl. Leaders; conditional RS excluded), keep
    # those whose RS line is within nh_tolerance of its 6-month high.
    # rs_line_tbl is indexed by Futu code, so map the TV survivors back to
    # codes for the lookup, then convert the selected codes to TV for output.
    import rs_line
    if rs_line.nh_is_enabled(config):
        nh_long_groups = ["EarningsGap", "HighVolume", "GapUp", "Leaders"]
        nh_candidate_codes = sorted({
            _tv_to_code(tv)
            for name in nh_long_groups
            for tv in final.get(name, [])
        })
        nh_tol = rs_line.nh_tolerance_from_config(config)
        nh_codes, nh_stats = rs_line.select_rs_new_high(
            nh_candidate_codes, rs_line_tbl, nh_tol
        )
        nh_tv = sorted(_to_tv(c) for c in nh_codes)
        dated_nh = hk_output_dir / f"{today_iso}_HKRSNewHigh.txt"
        write_watchlist(nh_tv, dated_nh, fmt)
        logger.info(
            f"[HK RS-NH] {rs_line.format_rs_new_high_summary(nh_stats)} "
            f"(tol={nh_tol:.0%}) -> {dated_nh}"
        )
        write_webull(nh_tv, dated_nh, output_dir)
        futu_sync(config, "hk_rs_new_high", nh_tv, "HK")
        tv_sync(config, "hk_rs_new_high", nh_tv, "HK")

    # --- HK IPO sidecar ---
    # Tickers with insufficient history for the IBD 12-month RS calc
    # (< 253 rows). Filtered through filter_hk_ipo_candidates (see its
    # docstring for the gate ladder). Independent cross-day master at
    # output/state/eod_seen_HKIPO.txt — a ticker collected today as IPO
    # still lands in its proper long-side group on the first day it has
    # enough history (the long-side master eod_seen_HK.txt is separate).
    #
    # 2026-05-21 tightening: added (1) hard minimum 20 trading days, and
    # (2) 3M RS >= min_rs_percentile_longs_3m at len(df) >= 64. See
    # filter_hk_ipo_candidates for the full ladder.
    ipo_codes, ipo_dropped = filter_hk_ipo_candidates(
        metrics, rs_table_3m, hk_settings
    )

    ipo_seen_path = eod_seen_path(output_dir, "HKIPO")
    ipo_seen = load_seen(ipo_seen_path)
    ipo_tv = sorted(_to_tv(c) for c in ipo_codes)
    rs_3m_threshold = int(hk_settings.get("min_rs_percentile_longs_3m", 90))
    logger.info(
        f"[HK IPO] {len(ipo_codes)} candidates after conditional filters "
        f"(>=20d hist; cap>={hk_settings.get('min_market_cap', 300_000_000):,.0f}, "
        f"price>={hk_settings.get('min_price', 20.0)}, "
        f"+if 20d: avg_vol/dvol/ADR, +if 50d: SMA50, +if 200d: SMA200, "
        f"+if 64d: RS_3M>={rs_3m_threshold}); "
        f"raw metrics<253: {int((metrics['n_rows'] < 253).sum())}; "
        f"dropped: {ipo_dropped}"
    )
    ipo_tv = dedup_seen("[HK IPO]", ipo_tv, ipo_seen, ipo_seen_path)
    dated_ipo = hk_output_dir / f"{today_iso}_IPO.txt"
    write_watchlist(ipo_tv, dated_ipo, fmt)
    logger.info(f"[HK IPO] {len(ipo_tv)} -> {dated_ipo}")
    write_webull(ipo_tv, dated_ipo, output_dir)
    futu_sync(config, "hk_ipo", ipo_tv, "HK")
    tv_sync(config, "hk_ipo", ipo_tv, "HK")

    # --- Per-category headline summary ---
    # One aligned line per group so the operator can see at a glance how many
    # each strategy produced and *why* a group is empty (RS-gated vs deduped),
    # without stitching together the per-stage funnel lines above.
    logger.info("[HK Longs] ── summary (found → RS-gate → within-dedup → written) ──")
    for n in HK_STRATEGY_PRIORITY:
        written = len(final.get(n, []))
        if n == "RS" and not rs_enabled:
            reason = "yesterday-close run" if use_yesterday else f"HSI day-change > {rs_trigger}"
            logger.info(f"[HK Longs]   {n:<11} skipped ({reason})")
            continue
        masked = len(dedup.get(n, [])) - written  # dropped by the cross-day master
        note = f"  [{masked} already-seen]" if masked > 0 else ""
        note += _rs_line_group_note(dedup.get(n, []), rs_line_tbl)
        logger.info(
            f"[HK Longs]   {n:<11} "
            f"{pre_counts.get(n, 0):>3} → {post_rs_counts.get(n, 0):>3} → "
            f"{len(dedup.get(n, [])):>3} → {written:>3}{note}"
        )
    ipo_masked = len(ipo_codes) - len(ipo_tv)
    ipo_note = f"  [{ipo_masked} already-seen]" if ipo_masked > 0 else ""
    logger.info(
        f"[HK Longs]   {'IPO':<11} {len(ipo_codes):>3} found"
        f"        → {len(ipo_tv):>3}{ipo_note}"
    )
